import openpyxl
import ipaddress

EXCEL_ROW_LIMIT = 1048576  # Define the Excel row limit

def expand_ip_range(start_ip, end_ip):
    start = list(map(int, start_ip.split(".")))
    end = list(map(int, end_ip.split(".")))
    temp = start
    ip_range = []

    ip_range.append(start_ip)
    while temp != end:
        start[3] += 1
        for i in (3, 2, 1):
            if temp[i] == 256:
                temp[i] = 0
                temp[i-1] += 1
        ip_range.append(".".join(map(str, temp)))
    return ip_range

with open('ip_addresses.txt', 'r') as file:
    ip_data = file.read().splitlines()

all_ips = []

for ip in ip_data:
    ip = ip.strip()  # remove leading/trailing whitespace
    if '-' in ip:
        try:
            start_ip, end_ip = map(str.strip, ip.split('-'))  # remove whitespace around IPs
            all_ips.extend(expand_ip_range(start_ip, end_ip))
        except ValueError:
            print(f"Skipping invalid range: {ip}")
    elif '/' in ip:
        try:
            ip = ip.rstrip(',')  # remove trailing comma if it exists
            all_ips.extend([str(ip) for ip in ipaddress.IPv4Network(ip, strict=False)])
        except ValueError:
            print(f"Skipping invalid CIDR range: {ip}")
    else:
        all_ips.append(ip)

wb = openpyxl.Workbook()
sheet = wb.active
sheet_counter = 1  # used for naming the sheets

row = 1  # Initialize row counter outside of loop

for ip in all_ips:
    if row > EXCEL_ROW_LIMIT:
        sheet_counter += 1
        sheet = wb.create_sheet(title=f"Sheet {sheet_counter}")
        row = 1  # Reset row counter for the new sheet
    sheet.cell(row=row, column=1, value=ip)
    row += 1

wb.save('output.xlsx')
